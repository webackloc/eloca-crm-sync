"""
listar_tabelas_bi.py — Inventário completo do banco BI ELOCA (biweback)
  - Lista todas as tabelas e views existentes com contagem de linhas
  - Mostra todas as colunas de cada tabela/view
  - Salva também um Excel com uma aba por tabela (primeiras 500 linhas de cada)

Uso: rodar via GitHub Actions (acesso ao BI liberado lá)
"""
import os
import sys
import pymssql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BI_HOST     = os.getenv("BI_HOST",     "og-bi.crwm94zs8mf9.sa-east-1.rds.amazonaws.com")
BI_PORT     = int(os.getenv("BI_PORT", "1433"))
BI_DATABASE = os.getenv("BI_DATABASE", "biweback")
BI_USER     = os.getenv("BI_USER",     "weback")
BI_PASSWORD = os.getenv("BI_PASSWORD", "")

OUTPUT_DIR  = "/tmp/bi_inventario"
OUTPUT_FILE = f"{OUTPUT_DIR}/bi_inventario_completo.xlsx"

os.makedirs(OUTPUT_DIR, exist_ok=True)

SEP = "═" * 70

print(f"\n{SEP}")
print(f"INVENTÁRIO BI ELOCA — banco: {BI_DATABASE}")
print(SEP)

conn = pymssql.connect(
    server=BI_HOST, port=BI_PORT, user=BI_USER,
    password=BI_PASSWORD, database=BI_DATABASE,
    timeout=120, charset="UTF-8",
)
cur = conn.cursor(as_dict=True)

# ── 1. Listar todas as tabelas e views ───────────────────────────────────────
cur.execute("""
    SELECT
        t.TABLE_NAME,
        t.TABLE_TYPE,
        SUM(p.rows) AS row_count
    FROM INFORMATION_SCHEMA.TABLES t
    LEFT JOIN sys.tables st   ON st.name = t.TABLE_NAME
    LEFT JOIN sys.partitions p ON p.object_id = st.object_id AND p.index_id IN (0,1)
    WHERE t.TABLE_SCHEMA = 'dbo'
    GROUP BY t.TABLE_NAME, t.TABLE_TYPE
    ORDER BY t.TABLE_TYPE, t.TABLE_NAME
""")
tabelas = cur.fetchall()

print(f"\nTotal encontrado: {len(tabelas)} objetos\n")
print(f"  {'NOME':<35} {'TIPO':<15} {'LINHAS':>10}")
print(f"  {'-'*35} {'-'*15} {'-'*10}")
for t in tabelas:
    rows = t['row_count'] if t['row_count'] else 0
    print(f"  {t['TABLE_NAME']:<35} {t['TABLE_TYPE']:<15} {rows:>10,}")

# ── 2. Colunas de cada tabela ────────────────────────────────────────────────
print(f"\n{SEP}")
print("COLUNAS POR TABELA/VIEW")
print(SEP)

cur.execute("""
    SELECT
        c.TABLE_NAME,
        c.COLUMN_NAME,
        c.DATA_TYPE,
        c.CHARACTER_MAXIMUM_LENGTH,
        c.IS_NULLABLE,
        c.ORDINAL_POSITION
    FROM INFORMATION_SCHEMA.COLUMNS c
    JOIN INFORMATION_SCHEMA.TABLES t
      ON t.TABLE_NAME = c.TABLE_NAME AND t.TABLE_SCHEMA = c.TABLE_SCHEMA
    WHERE c.TABLE_SCHEMA = 'dbo'
    ORDER BY c.TABLE_NAME, c.ORDINAL_POSITION
""")
colunas_raw = cur.fetchall()

# Agrupar por tabela
from collections import defaultdict
colunas_por_tabela = defaultdict(list)
for c in colunas_raw:
    colunas_por_tabela[c['TABLE_NAME']].append(c)

for tabela in sorted(colunas_por_tabela.keys()):
    cols = colunas_por_tabela[tabela]
    print(f"\n  {tabela} ({len(cols)} colunas):")
    for c in cols:
        size = f"({c['CHARACTER_MAXIMUM_LENGTH']})" if c['CHARACTER_MAXIMUM_LENGTH'] else ""
        null = "NULL" if c['IS_NULLABLE'] == 'YES' else "NOT NULL"
        print(f"    {c['COLUMN_NAME']:<30} {c['DATA_TYPE']}{size:<15} {null}")

# ── 3. Gerar Excel com amostra de cada tabela ────────────────────────────────
print(f"\n{SEP}")
print("GERANDO EXCEL com amostra de dados (até 500 linhas por tabela)...")
print(SEP)

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
DAT_FONT = Font(name="Arial", size=9)
ALT_FILL = PatternFill("solid", fgColor="EBF0FA")

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Aba 0: índice geral
ws_idx = wb.create_sheet("ÍNDICE")
ws_idx.append(["Tabela/View", "Tipo", "Linhas (aprox)", "Colunas"])
ws_idx.cell(1,1).font = HDR_FONT; ws_idx.cell(1,1).fill = HDR_FILL
ws_idx.cell(1,2).font = HDR_FONT; ws_idx.cell(1,2).fill = HDR_FILL
ws_idx.cell(1,3).font = HDR_FONT; ws_idx.cell(1,3).fill = HDR_FILL
ws_idx.cell(1,4).font = HDR_FONT; ws_idx.cell(1,4).fill = HDR_FILL

for t in tabelas:
    rows = t['row_count'] if t['row_count'] else 0
    ncols = len(colunas_por_tabela.get(t['TABLE_NAME'], []))
    ws_idx.append([t['TABLE_NAME'], t['TABLE_TYPE'], rows, ncols])

# Auto-largura índice
for col in ws_idx.columns:
    max_len = max((len(str(c.value or "")) for c in col), default=10)
    ws_idx.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)

# Aba por tabela
nomes_usados = set()
for t in tabelas:
    nome = t['TABLE_NAME']
    # Nome da aba (máx 31 chars, único)
    aba = nome[:31]
    if aba in nomes_usados:
        aba = nome[:28] + str(len(nomes_usados))
    nomes_usados.add(aba)

    try:
        cur.execute(f"SELECT TOP 500 * FROM [{nome}]")
        rows_data = cur.fetchall()
    except Exception as e:
        print(f"  AVISO: não foi possível ler {nome}: {e}")
        continue

    ws = wb.create_sheet(aba)
    if not rows_data:
        ws.append(["(sem dados)"])
        print(f"  {nome:<35} vazia")
        continue

    # Cabeçalho
    headers = list(rows_data[0].keys())
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, row in enumerate(rows_data, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row_idx, col_idx, row[key])
            cell.font = DAT_FONT
            cell.fill = fill

    # Auto-largura
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"  {nome:<35} {len(rows_data):>5} linhas amostradas")

conn.close()

wb.save(OUTPUT_FILE)
print(f"\n✓ Excel salvo em: {OUTPUT_FILE}")
print(f"  Abas geradas: {len(wb.sheetnames)}")
