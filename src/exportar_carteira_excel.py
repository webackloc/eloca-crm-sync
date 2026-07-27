"""
exportar_carteira_excel.py — Gera Excel completo de tudo que será enviado ao CRM

Sheets:
  1. Carteira_Contratos   — 133 contratos ativos com cliente, vigência, totais
  2. Equip_por_Contrato   — equipamentos ativos com serial, produto e valor unitário (via Supabase assets + ctprod)
  3. Produtos_Contrato    — itens ctprod com produto, descrição e valores
  4. Faturamento_2026     — notas fiscais emitidas em 2026

Uso (GitHub Actions — tem acesso ao BI e ao Supabase):
  BI_PASSWORD="..." SUPABASE_URL="..." SUPABASE_SERVICE_KEY="..." python3 src/exportar_carteira_excel.py
"""

import io
import os
import sys
import pymssql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuração de conexão ───────────────────────────────────────────────────
BI_HOST     = os.getenv("BI_HOST",     "og-bi.crwm94zs8mf9.sa-east-1.rds.amazonaws.com")
BI_PORT     = int(os.getenv("BI_PORT", "1433"))
BI_DATABASE = os.getenv("BI_DATABASE", "biweback")
BI_USER     = os.getenv("BI_USER",     "weback")
BI_PASSWORD = os.getenv("BI_PASSWORD", "")

SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")

OUTPUT_FILE = "/tmp/carteira_crm_validacao.xlsx"
OUTPUT_DIR  = "/tmp/excel_output"

# ── Estilos ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")      # azul escuro
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL  = PatternFill("solid", fgColor="EBF0FA")      # azul claro alternado
BRD       = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right =Side(style="thin", color="CCCCCC"),
)
CURRENCY_FMT = 'R$ #,##0.00'
DATE_FMT     = 'DD/MM/YYYY'

def apply_header(ws, headers: list[str]):
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BRD
    ws.row_dimensions[1].height = 28

def style_rows(ws, data_rows: int, currency_cols: list[int] = None, date_cols: list[int] = None):
    for row in range(2, data_rows + 2):
        fill = ALT_FILL if row % 2 == 0 else PatternFill()
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font   = DATA_FONT
            cell.fill   = fill
            cell.border = BRD
            cell.alignment = Alignment(vertical="center")
            if currency_cols and col in currency_cols:
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if date_cols and col in date_cols:
                cell.number_format = DATE_FMT
                cell.alignment = Alignment(horizontal="center", vertical="center")

def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ── Conexão BI ────────────────────────────────────────────────────────────────
print("Conectando ao BI SQL Server...")
conn = pymssql.connect(
    server=BI_HOST, port=BI_PORT, user=BI_USER,
    password=BI_PASSWORD, database=BI_DATABASE,
    timeout=120, charset="UTF-8",
)
cur = conn.cursor(as_dict=True)

# ── Busca assets do Supabase (serial + product_code por equipamento) ──────────
equip_assets = {}   # equip_code -> {"serial": ..., "product_code": ...}
seriais = {}        # mantido por compatibilidade
if SUPABASE_URL and SUPABASE_KEY:
    try:
        from supabase import create_client
        sb = create_client(SUPABASE_URL, SUPABASE_KEY)
        page, page_size = 0, 1000
        while True:
            res = (sb.table("assets")
                     .select("name,serial_number,product_code")
                     .range(page * page_size, (page + 1) * page_size - 1)
                     .execute())
            if not res.data:
                break
            for a in res.data:
                equip_code = str(a.get("name") or "")
                serial     = str(a.get("serial_number") or "")
                prod_code  = str(a.get("product_code") or "")
                equip_assets[equip_code] = {"serial": serial, "product_code": prod_code}
                if serial:
                    seriais[equip_code] = serial
            if len(res.data) < page_size:
                break
            page += 1
        print(f"  Assets carregados do Supabase (tabela assets): {len(equip_assets)}")
    except Exception as e:
        print(f"  Aviso: não foi possível carregar assets do Supabase: {e}")

# ═════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove sheet padrão

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Carteira_Contratos
# ══════════════════════════════════════════════════════════════════════════════
print("Gerando Sheet 1: Carteira_Contratos...")
ws1 = wb.create_sheet("Carteira_Contratos")

cur.execute("""
    WITH last_move AS (
        SELECT equipamento, contrato, envret,
               ROW_NUMBER() OVER (PARTITION BY equipamento ORDER BY data DESC, seq DESC) AS rn
        FROM ctmequip
    ),
    equip_ativos AS (
        SELECT contrato, COUNT(*) AS qtd_equip
        FROM last_move
        WHERE rn = 1 AND envret = 'E'
        GROUP BY contrato
    ),
    prod_contrato AS (
        SELECT contrato,
               COUNT(*) AS qtd_produtos,
               SUM(CONVERT(FLOAT, ISNULL(valor, 0))) AS valor_total
        FROM ctprod
        GROUP BY contrato
    ),
    fat_2026 AS (
        SELECT contrato,
               SUM(CONVERT(FLOAT, valoremissao)) AS total_faturado_2026
        FROM docrec
        WHERE LTRIM(RTRIM(ISNULL(contrato,''))) != ''
          AND dataemissao >= '2026-01-01'
        GROUP BY contrato
    )
    SELECT
        c.codigo                                                  AS contrato,
        (SELECT TOP 1 d.cliente FROM docrec d
         WHERE d.codigocliente = c.cliente
         ORDER BY d.recnum DESC)                                  AS cliente_nome,
        c.cliente                                                 AS cod_cliente,
        CONVERT(VARCHAR(10), c.datavigini, 120)                  AS data_inicio,
        CONVERT(VARCHAR(10), c.datavigfim, 120)                  AS data_fim,
        ISNULL(ea.qtd_equip, 0)                                  AS qtd_equip_ativos,
        ISNULL(pc.qtd_produtos, 0)                               AS qtd_produtos,
        ISNULL(pc.valor_total, 0)                                AS valor_total_mensal,
        ISNULL(f.total_faturado_2026, 0)                        AS faturado_2026
    FROM contract c
    LEFT JOIN equip_ativos   ea ON ea.contrato = c.codigo
    LEFT JOIN prod_contrato  pc ON pc.contrato = c.codigo
    LEFT JOIN fat_2026        f ON f.contrato  = c.codigo
    WHERE c.situacao = '3'
    ORDER BY CONVERT(INT, c.codigo)
""")
rows1 = cur.fetchall()

headers1 = ["Contrato", "Cliente", "Cód Cliente", "Início Vigência", "Fim Vigência",
            "Equip Ativos", "Qtd Produtos", "Valor Total Mensal (R$)", "Faturado 2026 (R$)"]
apply_header(ws1, headers1)

for r in rows1:
    ws1.append([
        r["contrato"], r["cliente_nome"] or "", r["cod_cliente"],
        r["data_inicio"], r["data_fim"],
        r["qtd_equip_ativos"], r["qtd_produtos"],
        float(r["valor_total_mensal"] or 0),
        float(r["faturado_2026"] or 0),
    ])

style_rows(ws1, len(rows1), currency_cols=[8, 9], date_cols=[4, 5])
# Totais
tot_row = len(rows1) + 2
ws1.cell(tot_row, 1, "TOTAL").font = Font(name="Arial", bold=True, size=10)
ws1.cell(tot_row, 6, f"=SUM(F2:F{tot_row-1})").font = Font(name="Arial", bold=True, size=10)
ws1.cell(tot_row, 8, f"=SUM(H2:H{tot_row-1})").number_format = CURRENCY_FMT
ws1.cell(tot_row, 8).font = Font(name="Arial", bold=True, size=10)
ws1.cell(tot_row, 9, f"=SUM(I2:I{tot_row-1})").number_format = CURRENCY_FMT
ws1.cell(tot_row, 9).font = Font(name="Arial", bold=True, size=10)

autofit(ws1)
freeze_and_filter(ws1)
print(f"  {len(rows1)} contratos")

# ── Lookup ctprod: (contrato, produto) -> valorunitario e descricao ───────────
print("Carregando lookup de produtos (ctprod)...")
cur.execute("""
    SELECT
        CONVERT(VARCHAR(20), cp.contrato) AS contrato,
        CONVERT(VARCHAR(20), cp.produto)  AS produto,
        ISNULL(CONVERT(VARCHAR(500), p.descricao), cp.produto) AS descricao,
        CONVERT(FLOAT, ISNULL(cp.valorunitario, 0)) AS valorunitario
    FROM ctprod cp
    LEFT JOIN produtos p ON p.codigo = cp.produto
    JOIN contract c ON c.codigo = cp.contrato
    WHERE c.situacao = '3'
""")
ctprod_lookup = {}
for _r in cur.fetchall():
    _key = (str(_r['contrato']), str(_r['produto']))
    ctprod_lookup[_key] = {
        'descricao': str(_r['descricao'] or ""),
        'valorunitario': float(_r['valorunitario'] or 0),
    }
print(f"  {len(ctprod_lookup)} combos contrato+produto carregados")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Equip_por_Contrato
# ══════════════════════════════════════════════════════════════════════════════
print("Gerando Sheet 2: Equip_por_Contrato...")
ws2 = wb.create_sheet("Equip_por_Contrato")

cur.execute("""
    WITH last_move AS (
        SELECT
            CONVERT(VARCHAR(20), recnum)     AS recnum,
            CONVERT(VARCHAR(20), equipamento) AS equipamento,
            CONVERT(VARCHAR(20), contrato)   AS contrato,
            CONVERT(VARCHAR(1),  envret)     AS envret,
            CONVERT(VARCHAR(10), data, 120)  AS data_movimento,
            ISNULL(CONVERT(VARCHAR(500), setor), '') AS setor,
            CONVERT(VARCHAR(20), numos)      AS numos,
            ROW_NUMBER() OVER (
                PARTITION BY equipamento
                ORDER BY data DESC, seq DESC
            ) AS rn
        FROM ctmequip
    )
    SELECT
        lm.contrato,
        (SELECT TOP 1 d.cliente FROM docrec d
         WHERE d.codigocliente = c.cliente
         ORDER BY d.recnum DESC) AS cliente_nome,
        lm.equipamento,
        lm.envret,
        lm.data_movimento,
        lm.setor,
        lm.numos,
        CONVERT(VARCHAR(10), c.datavigini, 120) AS contrato_inicio,
        CONVERT(VARCHAR(10), c.datavigfim, 120) AS contrato_fim
    FROM last_move lm
    JOIN contract c ON c.codigo = lm.contrato
    WHERE lm.rn = 1
      AND lm.envret = 'E'
      AND c.situacao = '3'
    ORDER BY CONVERT(INT, lm.contrato), lm.equipamento
""")
rows2 = cur.fetchall()

headers2 = ["Contrato", "Cliente", "Cód Equipamento", "Número Série",
            "Cód Produto", "Descrição Produto", "Valor Unitário (R$)",
            "Status", "Data Entrega", "Plano/Setor", "Nº OS",
            "Início Vigência", "Fim Vigência"]
apply_header(ws2, headers2)

for r in rows2:
    equip       = str(r["equipamento"] or "")
    contrato    = str(r["contrato"] or "")
    asset_info  = equip_assets.get(equip, {})
    serial      = asset_info.get("serial", "")
    prod_code   = asset_info.get("product_code", "")
    ctprod_info = ctprod_lookup.get((contrato, prod_code), {})
    descricao   = ctprod_info.get("descricao", "")
    valor_unit  = ctprod_info.get("valorunitario", None) if prod_code else None
    ws2.append([
        contrato, r["cliente_nome"] or "", equip, serial,
        prod_code, descricao, valor_unit,
        "ATIVO (entregue)", r["data_movimento"], r["setor"], r["numos"],
        r["contrato_inicio"], r["contrato_fim"],
    ])

style_rows(ws2, len(rows2), currency_cols=[7], date_cols=[9, 12, 13])
autofit(ws2)
freeze_and_filter(ws2)
print(f"  {len(rows2)} equipamentos ativos")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Produtos_Contrato
# ══════════════════════════════════════════════════════════════════════════════
print("Gerando Sheet 3: Produtos_Contrato...")
ws3 = wb.create_sheet("Produtos_Contrato")

cur.execute("""
    SELECT
        CONVERT(VARCHAR(20), cp.contrato)      AS contrato,
        (SELECT TOP 1 d.cliente FROM docrec d
         WHERE d.codigocliente = c.cliente
         ORDER BY d.recnum DESC)               AS cliente_nome,
        CONVERT(VARCHAR(20), cp.produto)       AS cod_produto,
        ISNULL(CONVERT(VARCHAR(500), p.descricao), cp.produto) AS descricao_produto,
        ISNULL(CONVERT(VARCHAR(500), cp.setor), '') AS segmento,
        CONVERT(FLOAT, ISNULL(cp.valorunitario, 0)) AS valor_unitario,
        CONVERT(FLOAT, ISNULL(cp.valor, 0))   AS valor_total_item
    FROM ctprod cp
    LEFT JOIN produtos p ON p.codigo = cp.produto
    JOIN contract c ON c.codigo = cp.contrato
    WHERE c.situacao = '3'
    ORDER BY CONVERT(INT, cp.contrato), cp.produto
""")
rows3 = cur.fetchall()

headers3 = ["Contrato", "Cliente", "Cód Produto", "Descrição do Produto",
            "Segmento/Plano", "Valor Unitário (R$)", "Valor Total Item (R$)"]
apply_header(ws3, headers3)

for r in rows3:
    ws3.append([
        r["contrato"], r["cliente_nome"] or "", r["cod_produto"],
        r["descricao_produto"], r["segmento"],
        float(r["valor_unitario"] or 0),
        float(r["valor_total_item"] or 0),
    ])

style_rows(ws3, len(rows3), currency_cols=[6, 7])
tot_row3 = len(rows3) + 2
ws3.cell(tot_row3, 1, "TOTAL").font = Font(name="Arial", bold=True, size=10)
ws3.cell(tot_row3, 7, f"=SUM(G2:G{tot_row3-1})").number_format = CURRENCY_FMT
ws3.cell(tot_row3, 7).font = Font(name="Arial", bold=True, size=10)
autofit(ws3)
freeze_and_filter(ws3)
print(f"  {len(rows3)} itens de produto")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Faturamento_2026
# ══════════════════════════════════════════════════════════════════════════════
print("Gerando Sheet 4: Faturamento_2026...")
ws4 = wb.create_sheet("Faturamento_2026")

cur.execute("""
    SELECT
        CONVERT(VARCHAR(30), d.numfatura)    AS numfatura,
        d.numsequencia                       AS parcela,
        ISNULL(NULLIF(LTRIM(RTRIM(CONVERT(VARCHAR(20), d.contrato))), ''), 'SEM CONTRATO') AS contrato,
        ISNULL(CONVERT(VARCHAR(200), d.cliente), '') AS cliente,
        CONVERT(FLOAT, d.valoremissao)       AS valor,
        CONVERT(VARCHAR(10), d.dataemissao, 120) AS data_emissao,
        CONVERT(VARCHAR(10), d.datavencto,  120) AS data_vencto,
        CASE WHEN LTRIM(RTRIM(ISNULL(d.liquidado,''))) = 'S'
             THEN 'SIM' ELSE 'NÃO' END      AS liquidado,
        ISNULL(CONVERT(VARCHAR(100), d.tipodocumento), '') AS tipo_documento,
        ISNULL(CONVERT(VARCHAR(200), c.representante_nome), '') AS vendedor
    FROM docrec d
    LEFT JOIN contract c ON c.codigo = d.contrato
    WHERE d.dataemissao >= '2026-01-01'
    ORDER BY d.dataemissao, d.numfatura
""")
rows4 = cur.fetchall()

headers4 = ["Nº Fatura", "Parcela", "Contrato", "Cliente",
            "Valor (R$)", "Data Emissão", "Data Vencimento",
            "Liquidado", "Tipo Documento", "Vendedor"]
apply_header(ws4, headers4)

for r in rows4:
    ws4.append([
        r["numfatura"], r["parcela"], r["contrato"], r["cliente"],
        float(r["valor"] or 0),
        r["data_emissao"], r["data_vencto"],
        r["liquidado"], r["tipo_documento"], r["vendedor"],
    ])

style_rows(ws4, len(rows4), currency_cols=[5], date_cols=[6, 7])
tot_row4 = len(rows4) + 2
ws4.cell(tot_row4, 1, "TOTAL").font = Font(name="Arial", bold=True, size=10)
ws4.cell(tot_row4, 5, f"=SUM(E2:E{tot_row4-1})").number_format = CURRENCY_FMT
ws4.cell(tot_row4, 5).font = Font(name="Arial", bold=True, size=10)
autofit(ws4)
freeze_and_filter(ws4)
print(f"  {len(rows4)} notas fiscais 2026")

conn.close()

# ── Salva o arquivo ───────────────────────────────────────────────────────────
import os as _os
_os.makedirs(OUTPUT_DIR, exist_ok=True)
dest = f"{OUTPUT_DIR}/carteira_crm_validacao.xlsx"
wb.save(dest)
print(f"\n✅ Excel salvo em: {dest}")
print("   Faça download na aba Actions → run atual → Artifacts → carteira-crm")
