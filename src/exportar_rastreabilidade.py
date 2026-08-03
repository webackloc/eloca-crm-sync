"""
exportar_rastreabilidade.py — View completa de rastreabilidade de equipamentos por contrato

Join:
  contract  → ctmequip  (contract.codigo = ctmequip.contrato)
  ctmequip  → equip     (ctmequip.equipamento = equip.codigo)
  equip     → ctprod    (equip.codigoproduto = ctprod.produto AND ctmequip.contrato = ctprod.contrato)

Inclui TODOS os movimentos (envios e retornos) para rastreabilidade histórica completa.
"""
import os
import pymssql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BI_HOST     = os.getenv("BI_HOST",     "og-bi.crwm94zs8mf9.sa-east-1.rds.amazonaws.com")
BI_PORT     = int(os.getenv("BI_PORT", "1433"))
BI_DATABASE = os.getenv("BI_DATABASE", "biweback")
BI_USER     = os.getenv("BI_USER",     "weback")
BI_PASSWORD = os.getenv("BI_PASSWORD", "")

OUTPUT_DIR  = "/tmp/rastreabilidade"
OUTPUT_FILE = f"{OUTPUT_DIR}/rastreabilidade_equipamentos.xlsx"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Estilos ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
DAT_FONT  = Font(name="Arial", size=9)
ALT_FILL  = PatternFill("solid", fgColor="EBF0FA")
ENV_FILL  = PatternFill("solid", fgColor="C6EFCE")   # verde — enviado
RET_FILL  = PatternFill("solid", fgColor="FFCCCC")   # vermelho — retorno
BRD = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right =Side(style="thin", color="CCCCCC"),
)
CURRENCY_FMT = 'R$ #,##0.00'
DATE_FMT     = 'DD/MM/YYYY'
DATETIME_FMT = 'DD/MM/YYYY HH:MM'

print("Conectando ao BI SQL Server...")
conn = pymssql.connect(
    server=BI_HOST, port=BI_PORT, user=BI_USER,
    password=BI_PASSWORD, database=BI_DATABASE,
    timeout=180, charset="UTF-8",
)
cur = conn.cursor(as_dict=True)

# ── Query principal ───────────────────────────────────────────────────────────
# ctprod pode ter múltiplas linhas por (contrato, produto) com valores diferentes
# → usamos o registro de maior recnum (mais recente) por (contrato, produto)
print("Executando query de rastreabilidade (pode levar alguns segundos)...")

cur.execute("""
    WITH ctprod_latest AS (
        SELECT
            CONVERT(VARCHAR(20), contrato) AS contrato,
            CONVERT(VARCHAR(20), produto)  AS produto,
            CONVERT(FLOAT, ISNULL(valorunitario, 0)) AS valorunitario,
            created_at,
            ROW_NUMBER() OVER (
                PARTITION BY contrato, produto
                ORDER BY recnum DESC
            ) AS rn
        FROM ctprod
    )
    SELECT
        -- ── contract ──────────────────────────────────────────────────────
        CONVERT(VARCHAR(20), c.codigo)                       AS contrato,
        CONVERT(VARCHAR(20), c.cliente)                      AS cod_cliente,
        ISNULL(CONVERT(VARCHAR(200), c.representante_nome), '') AS vendedor,
        CONVERT(VARCHAR(1),  c.situacao)                     AS situacao_contrato,
        CONVERT(VARCHAR(10), c.datavigini,    120)           AS datavigini,
        CONVERT(VARCHAR(10), c.datavigfim,    120)           AS datavigfim,
        CONVERT(VARCHAR(10), c.dataalteracao, 120)           AS dataalteracao,
        CONVERT(VARCHAR(19), c.created_at,   120)            AS contract_updated_bi,

        -- ── ctmequip ──────────────────────────────────────────────────────
        CONVERT(VARCHAR(20), m.equipamento)                  AS ativo,
        CONVERT(VARCHAR(1),  m.envret)                       AS envret,
        CONVERT(VARCHAR(10), m.data,          120)           AS data_movimento,
        ISNULL(CONVERT(VARCHAR(200), m.setor), '')           AS setor,
        CONVERT(VARCHAR(19), m.created_at,   120)            AS movimento_updated_bi,

        -- ── equip ─────────────────────────────────────────────────────────
        ISNULL(CONVERT(VARCHAR(20),  e.codigoproduto), '')   AS cod_produto,
        ISNULL(CONVERT(VARCHAR(500), e.produto), '')         AS descricao_produto,
        ISNULL(CONVERT(VARCHAR(200), e.seriefabricante), '') AS serial_fabricante,
        ISNULL(CONVERT(VARCHAR(50),  e.situacao), '')        AS situacao_equip,
        CONVERT(VARCHAR(19), e.created_at,   120)            AS equip_updated_bi,

        -- ── ctprod ────────────────────────────────────────────────────────
        ISNULL(cp.valorunitario, 0)                          AS valor_unitario,
        CONVERT(VARCHAR(19), cp.created_at,  120)            AS ctprod_updated_bi

    FROM ctmequip m
    JOIN contract c
      ON CONVERT(VARCHAR(20), c.codigo) = CONVERT(VARCHAR(20), m.contrato)
    LEFT JOIN equip e
      ON CONVERT(VARCHAR(20), e.codigo) = CONVERT(VARCHAR(20), m.equipamento)
    LEFT JOIN ctprod_latest cp
      ON cp.contrato = CONVERT(VARCHAR(20), m.contrato)
     AND cp.produto  = ISNULL(CONVERT(VARCHAR(20), e.codigoproduto), '')
     AND cp.rn = 1
    ORDER BY
        CONVERT(INT, c.codigo),
        m.equipamento,
        m.data DESC,
        m.seq  DESC
""")

rows = cur.fetchall()
conn.close()
print(f"  {len(rows):,} registros retornados")

# ── Situações legíveis ────────────────────────────────────────────────────────
SITUACAO_CONTRATO = {'3': 'APROVADO', '4': 'REPROVADO', '5': 'ENCERRADO'}

# ── Gerar Excel ───────────────────────────────────────────────────────────────
print("Gerando Excel...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 1 — Rastreabilidade completa (todos os movimentos)
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Rastreabilidade_Completa")

headers = [
    # contract
    "Contrato", "Cód Cliente", "Vendedor", "Situação Contrato",
    "Vigência Início", "Vigência Fim", "Última Alteração", "Atualiz. BI (contract)",
    # ctmequip
    "Ativo (Equipamento)", "E/R", "Data Movimento", "Setor",
    "Atualiz. BI (movimento)",
    # equip
    "Cód Produto", "Descrição Produto", "Serial Fabricante", "Situação Equip",
    "Atualiz. BI (equip)",
    # ctprod
    "Valor Unitário (R$)", "Atualiz. BI (ctprod)",
]

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(1, col_idx)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BRD
ws.row_dimensions[1].height = 32

for row_idx, r in enumerate(rows, 2):
    envret   = str(r['envret'] or '').strip()
    sit_cont = SITUACAO_CONTRATO.get(str(r['situacao_contrato'] or '').strip(), r['situacao_contrato'] or '')
    sit_eq   = str(r['situacao_equip'] or '').strip()
    vunit    = float(r['valor_unitario'] or 0)

    valores = [
        r['contrato'], r['cod_cliente'], r['vendedor'], sit_cont,
        r['datavigini'], r['datavigfim'], r['dataalteracao'], r['contract_updated_bi'],
        r['ativo'], 'ENVIADO' if envret == 'E' else 'RETORNO', r['data_movimento'], r['setor'],
        r['movimento_updated_bi'],
        r['cod_produto'], r['descricao_produto'], r['serial_fabricante'], sit_eq,
        r['equip_updated_bi'],
        vunit, r['ctprod_updated_bi'],
    ]

    ws.append(valores)

    # Cor por tipo de movimento
    row_fill = ENV_FILL if envret == 'E' else (RET_FILL if envret == 'R' else (ALT_FILL if row_idx % 2 == 0 else PatternFill()))

    for col_idx in range(1, len(headers) + 1):
        cell        = ws.cell(row_idx, col_idx)
        cell.font   = DAT_FONT
        cell.fill   = row_fill
        cell.border = BRD
        cell.alignment = Alignment(vertical="center")
        if col_idx == 19:   # Valor Unitário
            cell.number_format = CURRENCY_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")

# Auto-largura
for col in ws.columns:
    max_len = max((len(str(c.value or "")) for c in col), default=10)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
print(f"  Aba Rastreabilidade_Completa: {len(rows):,} linhas")

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 2 — Posição atual (último movimento por equipamento)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Posicao_Atual")

# Último movimento por equipamento
posicao = {}
for r in rows:
    key = str(r['ativo'] or '')
    if key not in posicao:
        posicao[key] = r   # já vem ordenado por data DESC

ws2.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws2.cell(1, col_idx)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BRD
ws2.row_dimensions[1].height = 32

ativos_enviados = [r for r in posicao.values() if str(r['envret'] or '').strip() == 'E']
em_retorno      = [r for r in posicao.values() if str(r['envret'] or '').strip() == 'R']
# Inconsistentes: INDISPONÍVEL mas último mov = R, ou situação diz disponível mas último = E
inconsistentes  = [r for r in posicao.values()
                   if ('INDISPON' in str(r['situacao_equip'] or '').upper() and str(r['envret'] or '').strip() == 'R')
                   or ('INDISPON' not in str(r['situacao_equip'] or '').upper() and str(r['envret'] or '').strip() == 'E')]

print(f"  Posição atual: {len(ativos_enviados)} enviados (E) / {len(em_retorno)} retornos (R)")
print(f"  Inconsistentes (sit_equip ≠ envret): {len(inconsistentes)}")

for row_idx, r in enumerate(sorted(posicao.values(), key=lambda x: (str(x['contrato']), str(x['ativo']))), 2):
    envret   = str(r['envret'] or '').strip()
    sit_eq   = str(r['situacao_equip'] or '').strip()
    sit_cont = SITUACAO_CONTRATO.get(str(r['situacao_contrato'] or '').strip(), r['situacao_contrato'] or '')
    vunit    = float(r['valor_unitario'] or 0)

    ws2.append([
        r['contrato'], r['cod_cliente'], r['vendedor'], sit_cont,
        r['datavigini'], r['datavigfim'], r['dataalteracao'], r['contract_updated_bi'],
        r['ativo'], 'ENVIADO' if envret == 'E' else 'RETORNO', r['data_movimento'], r['setor'],
        r['movimento_updated_bi'],
        r['cod_produto'], r['descricao_produto'], r['serial_fabricante'], sit_eq,
        r['equip_updated_bi'],
        vunit, r['ctprod_updated_bi'],
    ])

    row_fill = ENV_FILL if envret == 'E' else (RET_FILL if envret == 'R' else (ALT_FILL if row_idx % 2 == 0 else PatternFill()))
    for col_idx in range(1, len(headers) + 1):
        cell        = ws2.cell(row_idx, col_idx)
        cell.font   = DAT_FONT
        cell.fill   = row_fill
        cell.border = BRD
        cell.alignment = Alignment(vertical="center")
        if col_idx == 19:
            cell.number_format = CURRENCY_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")

for col in ws2.columns:
    max_len = max((len(str(c.value or "")) for c in col), default=10)
    ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions
print(f"  Aba Posicao_Atual: {len(posicao):,} equipamentos únicos")

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 3 — Resumo por contrato
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Resumo_por_Contrato")

from collections import defaultdict
resumo = defaultdict(lambda: {
    'vendedor': '', 'sit': '', 'vigini': '', 'vigfim': '',
    'total_mov': 0, 'ativos': 0, 'retornos': 0,
    'valor_carteira': 0.0, 'equips_unicos': set(),
})

for r in rows:
    c = str(r['contrato'])
    resumo[c]['vendedor']  = r['vendedor']
    resumo[c]['sit']       = SITUACAO_CONTRATO.get(str(r['situacao_contrato'] or '').strip(), '')
    resumo[c]['vigini']    = r['datavigini']
    resumo[c]['vigfim']    = r['datavigfim']
    resumo[c]['total_mov'] += 1
    resumo[c]['equips_unicos'].add(str(r['ativo'] or ''))
    envret = str(r['envret'] or '').strip()
    if envret == 'E':
        resumo[c]['ativos'] += 1
    elif envret == 'R':
        resumo[c]['retornos'] += 1

# Valor da carteira = equipamentos com último movimento ENVIADO (E) na posição atual
# Critério correto: envret='E' garante que o equip está fisicamente no cliente.
# Usar situacao_equip causava falsos positivos (INDISPONÍVEL + RETORNO = inconsistência BI).
for r in posicao.values():
    if str(r['envret'] or '').strip() == 'E':
        c = str(r['contrato'])
        resumo[c]['valor_carteira'] += float(r['valor_unitario'] or 0)

headers3 = ["Contrato", "Vendedor", "Situação", "Vigência Início", "Vigência Fim",
            "Total Movimentos", "Equip Únicos", "Ativos (E)", "Retornos (R)",
            "Valor Carteira (R$)"]
ws3.append(headers3)
for col_idx in range(1, len(headers3) + 1):
    cell = ws3.cell(1, col_idx)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = BRD
ws3.row_dimensions[1].height = 28

for row_idx, (contrato, d) in enumerate(sorted(resumo.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0), 2):
    ws3.append([
        contrato, d['vendedor'], d['sit'], d['vigini'], d['vigfim'],
        d['total_mov'], len(d['equips_unicos']),
        d['ativos'], d['retornos'],
        d['valor_carteira'],
    ])
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for col_idx in range(1, len(headers3) + 1):
        cell        = ws3.cell(row_idx, col_idx)
        cell.font   = DAT_FONT
        cell.fill   = fill
        cell.border = BRD
        cell.alignment = Alignment(vertical="center")
        if col_idx == 10:
            cell.number_format = CURRENCY_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")

# Total
tot = len(resumo) + 2
ws3.cell(tot, 1, "TOTAL").font = Font(name="Arial", bold=True, size=9)
ws3.cell(tot, 6, f"=SUM(F2:F{tot-1})").font = Font(name="Arial", bold=True, size=9)
ws3.cell(tot, 7, f"=SUM(G2:G{tot-1})").font = Font(name="Arial", bold=True, size=9)
ws3.cell(tot, 8, f"=SUM(H2:H{tot-1})").font = Font(name="Arial", bold=True, size=9)
ws3.cell(tot, 9, f"=SUM(I2:I{tot-1})").font = Font(name="Arial", bold=True, size=9)
ws3.cell(tot, 10, f"=SUM(J2:J{tot-1})").number_format = CURRENCY_FMT
ws3.cell(tot, 10).font = Font(name="Arial", bold=True, size=9)

for col in ws3.columns:
    max_len = max((len(str(c.value or "")) for c in col), default=10)
    ws3.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)

ws3.freeze_panes = "A2"
ws3.auto_filter.ref = ws3.dimensions
print(f"  Aba Resumo_por_Contrato: {len(resumo)} contratos")

# ── Salvar ────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\n✓ Excel salvo: {OUTPUT_FILE}")
print(f"  Abas: {wb.sheetnames}")
