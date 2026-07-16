"""
export_bi_excel.py — Exporta todas as tabelas do BI para Excel (uma aba por tabela)
Uso: python3 src/export_bi_excel.py
Saída: bi_export_YYYYMMDD.xlsx na pasta atual
"""
import os
import sys
from datetime import datetime

try:
    import pymssql
except ImportError:
    print("Instalando pymssql...")
    os.system(f"{sys.executable} -m pip install pymssql --break-system-packages -q")
    import pymssql

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# ── Conexão BI ────────────────────────────────────────────────────────────────
conn = pymssql.connect(
    server   = os.getenv("BI_HOST", "og-bi.crwm94zs8mf9.sa-east-1.rds.amazonaws.com"),
    port     = int(os.getenv("BI_PORT", "1433")),
    user     = os.getenv("BI_USER", "weback"),
    password = os.getenv("BI_PASSWORD", ""),
    database = os.getenv("BI_DATABASE", "biweback"),
    timeout  = 120, charset="UTF-8",
)
cur = conn.cursor(as_dict=True)

# Descobre tabelas
cur2 = conn.cursor()
cur2.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' ORDER BY TABLE_NAME")
tabelas = [r[0] for r in cur2.fetchall()]
print(f"Tabelas encontradas: {tabelas}")

# ── Excel ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove aba padrão

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")

for tabela in tabelas:
    print(f"Exportando {tabela}...", end=" ")
    cur.execute(f"SELECT * FROM {tabela}")
    rows = cur.fetchall()
    print(f"{len(rows):,} registros")

    ws = wb.create_sheet(title=tabela[:31])  # Excel limita nome da aba a 31 chars

    if not rows:
        ws.append(["(sem dados)"])
        continue

    # Cabeçalho
    colunas = list(rows[0].keys())
    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(colunas, 1):
            val = row[col_name]
            # Converte tipos não serializáveis
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill

    # Ajusta largura das colunas
    for col_idx, col_name in enumerate(colunas, 1):
        max_len = max(len(str(col_name)), 10)
        for row in rows[:50]:  # amostra das primeiras 50 linhas
            v = row[col_name]
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Congela cabeçalho
    ws.freeze_panes = "A2"

conn.close()

# Salva
nome = f"bi_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
saida = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), nome)
wb.save(saida)
print(f"\n✓ Exportado: {saida}")
